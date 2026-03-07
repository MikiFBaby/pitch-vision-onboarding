#!/usr/bin/env python3
"""
Import payroll data from American + Canadian XLSX files into Supabase.
Populates the payroll_periods table and fills missing hourly_wage in employee_directory.

Usage:
    python scripts/payroll-import.py --dry-run
    python scripts/payroll-import.py
    python scripts/payroll-import.py --american ~/Desktop/file.xlsx --canadian ~/Desktop/file.xlsx
    python scripts/payroll-import.py --audit-only
"""

import os
import sys
import json
import re
import argparse
from difflib import SequenceMatcher

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: pip3 install requests")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
BLOCKLIST_PATH = os.path.join(os.path.dirname(__file__), "pitch-health-blocklist.json")

DEFAULT_AMERICAN_FILE = os.path.expanduser(
    "~/Desktop/American Pay February 2nd - February 14th, 2026 2.xlsx"
)
DEFAULT_CANADIAN_FILE = os.path.expanduser(
    "~/Desktop/Canadian Pay February 8th - February 21st, 2026 1.xlsx"
)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

HEADERS_UPSERT = {
    **HEADERS,
    "Prefer": "resolution=merge-duplicates,return=minimal",
}

# Month name → number mapping
MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def load_blocklist():
    if os.path.exists(BLOCKLIST_PATH):
        with open(BLOCKLIST_PATH) as f:
            data = json.load(f)
        return set(n.strip().lower() for n in data)
    return set()


def load_env():
    """Load .env.local if SUPABASE_URL not set."""
    global SUPABASE_URL, SUPABASE_KEY, HEADERS, HEADERS_UPSERT
    if SUPABASE_URL and SUPABASE_KEY:
        return
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env.local")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip().strip('"').strip("'")
        SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
        SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
        HEADERS["apikey"] = SUPABASE_KEY
        HEADERS["Authorization"] = f"Bearer {SUPABASE_KEY}"
        HEADERS_UPSERT["apikey"] = SUPABASE_KEY
        HEADERS_UPSERT["Authorization"] = f"Bearer {SUPABASE_KEY}"


def fetch_employees():
    """Fetch active employees from employee_directory."""
    url = f"{SUPABASE_URL}/rest/v1/employee_directory?select=id,first_name,last_name,hourly_wage,country,employee_status,dialedin_name&employee_status=eq.Active"
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()
    return resp.json()


def safe_float(val):
    """Convert cell value to float, handling None, #N/A, strings."""
    if val is None:
        return 0.0
    if isinstance(val, str):
        val = val.strip().replace("$", "").replace(",", "")
        if val in ("", "#N/A", "N/A", "-"):
            return 0.0
        try:
            return float(val)
        except ValueError:
            return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def safe_int(val):
    return int(safe_float(val))


def parse_date_from_sheet_name(sheet_name):
    """
    Extract date range from sheet name like "February 1st - February 7th" or "January 18th - January 24th".
    Returns (start_date, end_date) as YYYY-MM-DD strings, or (None, None) if not parseable.
    """
    # Pattern: "Month Dayth - Month Dayth" (with optional ordinal suffixes)
    pattern = r'(\w+)\s+(\d+)\w*\s*-\s*(\w+)\s+(\d+)\w*'
    m = re.search(pattern, sheet_name)
    if not m:
        return None, None

    start_month_name = m.group(1).lower()
    start_day = int(m.group(2))
    end_month_name = m.group(3).lower()
    end_day = int(m.group(4))

    start_month = MONTH_MAP.get(start_month_name)
    end_month = MONTH_MAP.get(end_month_name)
    if not start_month or not end_month:
        return None, None

    # Assume 2026 (current year)
    year = 2026
    # Handle year boundary (Dec → Jan)
    start_year = year if start_month <= 12 else year
    end_year = year
    if start_month == 12 and end_month == 1:
        end_year = year + 1

    start_date = f"{start_year}-{start_month:02d}-{start_day:02d}"
    end_date = f"{end_year}-{end_month:02d}-{end_day:02d}"
    return start_date, end_date


def parse_weekly_sheet(ws, country, period_start, period_end):
    """Parse a weekly payroll sheet (row 2 = headers, row 3+ = data)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return []

    # Row 0 is headers
    header = [str(c).strip() if c else "" for c in rows[0]]

    # Map column indices
    col_map = {}
    for i, h in enumerate(header):
        hl = h.lower()
        if "employee name" in hl:
            col_map["name"] = i
        elif hl == "total hours":
            col_map["hours"] = i
        elif "hourly rate" in hl or "hourly wage" in hl:
            col_map["rate"] = i
        elif "total hourly pay" in hl:
            col_map["hourly_pay"] = i
        elif hl == "sla":
            col_map["sla"] = i
        elif "total transfers" in hl:
            col_map["transfers"] = i
        elif "commis" in hl and "total" in hl:
            col_map["commission"] = i
        elif hl == "bonus":
            col_map["bonus"] = i
        elif hl == "total pay":
            col_map["total_pay"] = i

    if "name" not in col_map:
        print(f"    WARNING: No 'Employee Name' column found in sheet")
        return []

    agents = []
    for row in rows[1:]:  # Skip header
        if col_map["name"] >= len(row):
            continue
        name_val = row[col_map["name"]]
        if not name_val or not isinstance(name_val, str):
            continue
        name = name_val.strip()
        if not name or name.upper() in ("GRAND TOTAL", "TOTAL"):
            continue

        def get_val(key):
            idx = col_map.get(key)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        agents.append({
            "agent_name": name,
            "period_start": period_start,
            "period_end": period_end,
            "country": country,
            "hours_worked": safe_float(get_val("hours")),
            "hourly_rate": safe_float(get_val("rate")),
            "hourly_pay": safe_float(get_val("hourly_pay")),
            "sla_transfers": safe_int(get_val("transfers")),
            "commission": safe_float(get_val("commission")),
            "bonus": safe_float(get_val("bonus")),
            "total_pay": safe_float(get_val("total_pay")),
        })

    return agents


def parse_grand_total_sheet(ws, country):
    """Parse a Grand Total sheet for wages only (cross-reference/audit)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c else "" for c in rows[0]]

    col_map = {}
    for i, h in enumerate(header):
        hl = h.lower().strip()
        if "employee name" in hl:
            col_map["name"] = i
        elif "hourly" in hl and ("rate" in hl or "wage" in hl):
            col_map["rate"] = i
        elif "total hours" in hl:
            col_map["hours"] = i
        elif "total pay" in hl:
            col_map["total_pay"] = i

    if "name" not in col_map:
        return []

    results = []
    for row in rows[1:]:
        if col_map["name"] >= len(row):
            continue
        name_val = row[col_map["name"]]
        if not name_val or not isinstance(name_val, str):
            continue
        name = name_val.strip()
        if not name or name.upper() in ("GRAND TOTAL", "TOTAL"):
            continue

        rate_idx = col_map.get("rate")
        rate = safe_float(row[rate_idx] if rate_idx is not None and rate_idx < len(row) else None)
        hours_idx = col_map.get("hours")
        hours = safe_float(row[hours_idx] if hours_idx is not None and hours_idx < len(row) else None)
        pay_idx = col_map.get("total_pay")
        total_pay = safe_float(row[pay_idx] if pay_idx is not None and pay_idx < len(row) else None)

        results.append({
            "agent_name": name,
            "country": country,
            "hourly_rate": rate,
            "hours_worked": hours,
            "total_pay": total_pay,
        })

    return results


def parse_workbook(filepath, country):
    """Auto-iterate all sheets in a workbook. Returns (weekly_agents, grand_total_agents)."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    weekly_agents = []
    grand_total_agents = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip "Sheet2" and other utility sheets
        if sheet_name.lower().startswith("sheet"):
            continue

        # Grand Total sheet — parse for wage cross-reference
        if "grand total" in sheet_name.lower():
            agents = parse_grand_total_sheet(ws, country)
            grand_total_agents.extend(agents)
            print(f"  Grand Total: {len(agents)} agents")
            continue

        # Weekly sheet — try to extract dates from name
        start_date, end_date = parse_date_from_sheet_name(sheet_name)
        if start_date and end_date:
            agents = parse_weekly_sheet(ws, country, start_date, end_date)
            print(f"  Sheet \"{sheet_name}\" ({start_date} → {end_date}): {len(agents)} agents")
            weekly_agents.extend(agents)
        else:
            print(f"  Skipping sheet \"{sheet_name}\" (could not parse dates)")

    wb.close()
    return weekly_agents, grand_total_agents


def build_name_lookup(employees):
    """Build tiered name lookup from employee_directory."""
    exact = {}  # "first last" → employee
    last_initial = {}  # "last f" → employee

    for emp in employees:
        fn = (emp.get("first_name") or "").strip()
        ln = (emp.get("last_name") or "").strip()
        full = f"{fn} {ln}".strip().lower()
        if full:
            exact[full] = emp

        if ln and fn:
            key = f"{ln} {fn[0]}".lower()
            last_initial[key] = emp

    return exact, last_initial


def match_employee(name, exact_map, last_initial_map, all_employees):
    """Tiered name matching: exact → last+initial → fuzzy."""
    name_lower = name.strip().lower()

    # Tier 1: Exact match
    if name_lower in exact_map:
        return exact_map[name_lower], "exact"

    # Tier 2: Last name + first initial
    parts = name_lower.split()
    if len(parts) >= 2:
        key = f"{parts[-1]} {parts[0][0]}"
        if key in last_initial_map:
            return last_initial_map[key], "last_initial"

    # Tier 3: Fuzzy match
    best_score = 0
    best_emp = None
    for emp in all_employees:
        fn = (emp.get("first_name") or "").strip()
        ln = (emp.get("last_name") or "").strip()
        full = f"{fn} {ln}".strip().lower()
        if not full:
            continue
        score = SequenceMatcher(None, name_lower, full).ratio()
        if score > best_score:
            best_score = score
            best_emp = emp
    if best_score >= 0.8:
        return best_emp, "fuzzy"

    return None, None


def main():
    parser = argparse.ArgumentParser(description="Import payroll XLSX to Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Parse and display without writing to DB")
    parser.add_argument("--audit-only", action="store_true", help="Cross-reference only, no DB writes")
    parser.add_argument("--american", type=str, help="Path to American payroll XLSX")
    parser.add_argument("--canadian", type=str, help="Path to Canadian payroll XLSX")
    args = parser.parse_args()

    load_env()
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set")
        sys.exit(1)

    american_file = os.path.expanduser(args.american) if args.american else DEFAULT_AMERICAN_FILE
    canadian_file = os.path.expanduser(args.canadian) if args.canadian else DEFAULT_CANADIAN_FILE

    blocklist = load_blocklist()
    print(f"Loaded Pitch Health blocklist: {len(blocklist)} names")

    # ── Parse payroll files ────────────────────────────────────
    all_weekly = []
    all_grand_total = []

    if os.path.exists(american_file):
        print(f"\nParsing American payroll: {os.path.basename(american_file)}")
        weekly, grand = parse_workbook(american_file, "USA")
        all_weekly.extend(weekly)
        all_grand_total.extend(grand)
    else:
        print(f"WARNING: American payroll file not found: {american_file}")

    if os.path.exists(canadian_file):
        print(f"\nParsing Canadian payroll: {os.path.basename(canadian_file)}")
        weekly, grand = parse_workbook(canadian_file, "Canada")
        all_weekly.extend(weekly)
        all_grand_total.extend(grand)
    else:
        print(f"WARNING: Canadian payroll file not found: {canadian_file}")

    # ── Filter Pitch Health ───────────────────────────────────
    before_weekly = len(all_weekly)
    all_weekly = [a for a in all_weekly if a["agent_name"].strip().lower() not in blocklist]
    before_grand = len(all_grand_total)
    all_grand_total = [a for a in all_grand_total if a["agent_name"].strip().lower() not in blocklist]
    filtered = (before_weekly - len(all_weekly)) + (before_grand - len(all_grand_total))
    print(f"\nFiltered {filtered} Pitch Health agents")
    print(f"  Weekly records: {len(all_weekly)}, Grand Total records: {len(all_grand_total)}")

    # ── Cross-reference with employee_directory ───────────────
    print("\nFetching employee directory...")
    employees = fetch_employees()
    print(f"  {len(employees)} Active employees")

    exact_map, li_map = build_name_lookup(employees)

    # Match weekly records
    matched = 0
    unmatched = []
    wage_updates = []

    # Build unique agent→wage from grand total (most reliable wage source)
    gt_wages = {}
    for gt in all_grand_total:
        if gt["hourly_rate"] > 0:
            gt_wages[gt["agent_name"].strip().lower()] = gt["hourly_rate"]

    for agent in all_weekly:
        emp, method = match_employee(agent["agent_name"], exact_map, li_map, employees)
        if emp:
            agent["employee_id"] = emp["id"]
            matched += 1

            # Check for missing wages — prefer Grand Total rate, fall back to weekly
            rate = gt_wages.get(agent["agent_name"].strip().lower(), agent["hourly_rate"])
            if emp.get("hourly_wage") is None and rate > 0:
                wage_updates.append({
                    "id": emp["id"],
                    "name": f"{emp['first_name']} {emp['last_name']}",
                    "hourly_wage": rate,
                })

            # Check for missing dialedin_name — payroll names are close to directory names
            # but we don't auto-set dialedin_name from payroll (it comes from DialedIn)
        else:
            agent["employee_id"] = None
            if agent["agent_name"] not in unmatched:
                unmatched.append(agent["agent_name"])

    # Deduplicate wage updates (same employee may appear in multiple weeks)
    seen_wage_ids = set()
    deduped_wages = []
    for w in wage_updates:
        if w["id"] not in seen_wage_ids:
            seen_wage_ids.add(w["id"])
            deduped_wages.append(w)
    wage_updates = deduped_wages

    print(f"  Matched: {matched} records | Unmatched agents: {len(unmatched)}")
    if wage_updates:
        print(f"  Missing wages to fill: {len(wage_updates)}")
        for w in wage_updates:
            print(f"    {w['name']}: ${w['hourly_wage']}")

    if unmatched[:10]:
        print(f"\n  First 10 unmatched: {', '.join(unmatched[:10])}")

    # ── Wage discrepancy audit ────────────────────────────────
    print("\n── Wage Discrepancy Check ──")
    discrepancies = []
    for gt in all_grand_total:
        emp, method = match_employee(gt["agent_name"], exact_map, li_map, employees)
        if emp and emp.get("hourly_wage") is not None and gt["hourly_rate"] > 0:
            db_wage = float(emp["hourly_wage"])
            payroll_wage = gt["hourly_rate"]
            diff = abs(db_wage - payroll_wage)
            if diff > 0.50:
                discrepancies.append({
                    "name": f"{emp['first_name']} {emp['last_name']}",
                    "db_wage": db_wage,
                    "payroll_wage": payroll_wage,
                    "diff": diff,
                    "country": gt["country"],
                })

    if discrepancies:
        print(f"  Found {len(discrepancies)} wage discrepancies (> $0.50 difference):")
        for d in sorted(discrepancies, key=lambda x: x["diff"], reverse=True):
            print(f"    {d['name']}: DB ${d['db_wage']:.2f} vs Payroll ${d['payroll_wage']:.2f} (diff ${d['diff']:.2f}) [{d['country']}]")
    else:
        print("  No wage discrepancies found")

    # ── Suspicious wages ──────────────────────────────────────
    suspicious = [gt for gt in all_grand_total if gt["hourly_rate"] > 0 and (gt["hourly_rate"] < 12 or gt["hourly_rate"] > 35)]
    if suspicious:
        print(f"\n  Suspicious wages (< $12 or > $35):")
        for s in suspicious:
            print(f"    {s['agent_name']}: ${s['hourly_rate']:.2f} [{s['country']}]")

    # ── Summary ────────────────────────────────────────────────
    usa_weekly = [a for a in all_weekly if a["country"] == "USA"]
    can_weekly = [a for a in all_weekly if a["country"] == "Canada"]
    print(f"\nSummary:")
    print(f"  USA weekly records: {len(usa_weekly)}, total pay: ${sum(a['total_pay'] for a in usa_weekly):,.2f}")
    print(f"  Canada weekly records: {len(can_weekly)}, total pay: ${sum(a['total_pay'] for a in can_weekly):,.2f}")
    print(f"  Grand Total agents: USA={len([g for g in all_grand_total if g['country']=='USA'])}, Canada={len([g for g in all_grand_total if g['country']=='Canada'])}")

    if args.audit_only:
        print(f"\n[AUDIT ONLY] No DB writes performed")
        return

    if args.dry_run:
        print(f"\n[DRY RUN] Would insert {len(all_weekly)} payroll records")
        print(f"[DRY RUN] Would update {len(wage_updates)} missing wages")
        return

    # ── Write to payroll_periods ──────────────────────────────
    print(f"\nInserting {len(all_weekly)} payroll records...")
    url = f"{SUPABASE_URL}/rest/v1/payroll_periods"

    # Batch upsert in chunks of 100
    batch_size = 100
    inserted = 0
    for i in range(0, len(all_weekly), batch_size):
        batch = all_weekly[i : i + batch_size]
        resp = requests.post(url, headers=HEADERS_UPSERT, json=batch)
        if resp.status_code in (200, 201):
            inserted += len(batch)
            print(f"  Inserted {inserted}/{len(all_weekly)}")
        else:
            print(f"  ERROR at batch {i}: {resp.status_code} {resp.text[:200]}")

    # ── Fill missing wages ────────────────────────────────────
    if wage_updates:
        print(f"\nUpdating {len(wage_updates)} missing hourly wages...")
        for w in wage_updates:
            url = f"{SUPABASE_URL}/rest/v1/employee_directory?id=eq.{w['id']}"
            resp = requests.patch(url, headers=HEADERS, json={"hourly_wage": w["hourly_wage"]})
            if resp.status_code in (200, 204):
                print(f"  Updated {w['name']}: ${w['hourly_wage']}")
            else:
                print(f"  ERROR updating {w['name']}: {resp.status_code}")

    print("\nDone!")


if __name__ == "__main__":
    main()
